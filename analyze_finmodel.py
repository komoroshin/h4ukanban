import pandas as pd
import openpyxl
from openpyxl import load_workbook

# Путь к файлу
file_path = 'finmodel/Office Open XML spreadsheet.xlsx'

# Загружаем книгу
wb = load_workbook(file_path, data_only=True)

print("=" * 80)
print("АНАЛИЗ ФИНАНСОВОЙ МОДЕЛИ")
print("=" * 80)
print(f"\nНазвание файла: {file_path}")
print(f"\nЛисты в книге: {wb.sheetnames}")
print("=" * 80)

# Анализируем каждый лист
for sheet_name in wb.sheetnames:
    print(f"\n\n{'=' * 80}")
    print(f"ЛИСТ: {sheet_name}")
    print("=" * 80)

    # Читаем данные с помощью pandas
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

    print(f"\nРазмер данных: {df.shape[0]} строк x {df.shape[1]} столбцов")
    print("\nПервые 30 строк:")
    print(df.head(30).to_string())

    # Проверяем наличие пустых значений
    if df.isna().any().any():
        print("\n⚠️  ВНИМАНИЕ: Обнаружены пустые ячейки")
        print("\nПустые значения по столбцам:")
        print(df.isna().sum())

print("\n" + "=" * 80)
print("АНАЛИЗ ЗАВЕРШЕН")
print("=" * 80)
