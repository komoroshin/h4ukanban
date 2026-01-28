import pandas as pd
from openpyxl import load_workbook

file_path = 'finmodel/Office Open XML spreadsheet.xlsx'
wb = load_workbook(file_path, data_only=True)

print("=" * 80)
print("СТРУКТУРА ФИНАНСОВОЙ МОДЕЛИ")
print("=" * 80)

sheets_info = []
issues = []

for sheet_name in wb.sheetnames:
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

    # Базовая информация
    rows, cols = df.shape
    non_empty = df.notna().sum().sum()
    empty_cells = rows * cols - non_empty

    # Пытаемся определить назначение листа
    first_cells = df.iloc[:5, :5].fillna('').astype(str).values.flatten()
    content_preview = ' | '.join([c[:20] for c in first_cells if c][:3])

    sheets_info.append({
        'Лист': sheet_name,
        'Размер': f"{rows}x{cols}",
        'Заполнено': f"{non_empty}/{rows*cols}",
        'Пустых': empty_cells,
        'Превью': content_preview[:60]
    })

    # Проверка на проблемы
    if empty_cells > (rows * cols * 0.9):
        issues.append(f"⚠️  {sheet_name}: Более 90% пустых ячеек")

    # Проверяем наличие ошибок в формулах
    ws = wb[sheet_name]
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if '#' in str(cell.value) and any(err in str(cell.value) for err in ['#REF!', '#VALUE!', '#DIV/0!', '#N/A', '#NAME?']):
                    issues.append(f"❌ {sheet_name} [{cell.coordinate}]: Ошибка в формуле - {cell.value}")

# Вывод результатов
print("\nЛисты в модели:")
print("-" * 80)
for info in sheets_info:
    print(f"\n📊 {info['Лист']}")
    print(f"   Размер: {info['Размер']} | Заполнено: {info['Заполнено']}")
    print(f"   Содержание: {info['Превью']}")

print("\n" + "=" * 80)
print("ОБНАРУЖЕННЫЕ ПРОБЛЕМЫ")
print("=" * 80)

if issues:
    for issue in issues:
        print(issue)
else:
    print("✅ Критических проблем не обнаружено")

print("\n" + "=" * 80)
print("РЕКОМЕНДАЦИИ")
print("=" * 80)

# Проверяем наличие ключевых листов
expected_sheets = {
    'Assumptions': 'Базовые предположения и допущения',
    'Revenue Model': 'Модель доходов',
    'Cost Structure': 'Структура расходов',
    'Cash Flow': 'Денежный поток',
    'P&L': 'Прибыли и убытки',
    'Balance Sheet': 'Баланс',
    'KPI': 'Ключевые показатели'
}

print("\nПроверка структуры модели:")
for expected, description in expected_sheets.items():
    found = any(expected.lower() in sheet.lower() for sheet in wb.sheetnames)
    status = "✅" if found else "❓"
    print(f"{status} {expected} - {description}")

print("\n" + "=" * 80)
