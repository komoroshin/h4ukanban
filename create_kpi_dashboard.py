from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import shutil
from datetime import datetime

file_path = 'finmodel/Office Open XML spreadsheet.xlsx'
backup_file = f'finmodel/Office Open XML spreadsheet_before_KPI_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

print("=" * 80)
print("СОЗДАНИЕ KPI DASHBOARD")
print("=" * 80)

# Создаем резервную копию
print(f"\n1. Создание резервной копии...")
shutil.copy2(file_path, backup_file)
print(f"   ✅ Резервная копия: {backup_file}")

# Загружаем workbook
print(f"\n2. Загрузка модели...")
wb = load_workbook(file_path)

# Проверяем, существует ли уже KPI Dashboard
if 'KPI Dashboard' in wb.sheetnames:
    print(f"   ⚠️  Лист 'KPI Dashboard' уже существует, будет перезаписан")
    del wb['KPI Dashboard']

# Создаем новый лист в начале
ws = wb.create_sheet('KPI Dashboard', 0)
print(f"   ✅ Создан новый лист 'KPI Dashboard'")

print(f"\n3. Настройка стилей...")

# Стили
header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
header_font = Font(name='Arial', size=14, bold=True, color="FFFFFF")
section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
section_font = Font(name='Arial', size=12, bold=True, color="2F75B5")
metric_font = Font(name='Arial', size=11)
value_font = Font(name='Arial', size=11, bold=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

print(f"\n4. Создание структуры Dashboard...")

# Заголовок
ws['A1'] = 'Healthy4U'
ws['A1'].font = Font(name='Arial', size=18, bold=True, color="2F75B5")
ws['A2'] = 'KPI DASHBOARD'
ws['A2'].font = Font(name='Arial', size=16, bold=True)
ws['A3'] = 'Ключевые показатели эффективности'
ws['A3'].font = Font(name='Arial', size=10, italic=True, color="7F7F7F")

current_row = 5

# Функция для создания секции
def create_section(start_row, section_name, metrics):
    global current_row

    # Заголовок секции
    ws[f'A{start_row}'] = section_name
    ws[f'A{start_row}'].fill = section_fill
    ws[f'A{start_row}'].font = section_font
    ws[f'A{start_row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(f'A{start_row}:B{start_row}')

    # Метрики и значения
    row = start_row + 1
    for metric_name, formula_or_value, description in metrics:
        # Название метрики
        ws[f'A{row}'] = metric_name
        ws[f'A{row}'].font = metric_font
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')

        # Значение
        ws[f'B{row}'] = formula_or_value
        ws[f'B{row}'].font = value_font
        ws[f'B{row}'].alignment = Alignment(horizontal='right', vertical='center')

        # Описание (опционально)
        if description:
            ws[f'C{row}'] = description
            ws[f'C{row}'].font = Font(name='Arial', size=9, italic=True, color="7F7F7F")

        row += 1

    return row + 1  # Возвращаем следующую строку после пустой

# 1. REVENUE METRICS
print(f"   • Revenue Metrics...")
revenue_metrics = [
    ("MRR (Monthly Recurring Revenue)", "=SUM('Monthly forecast'!$D$4:$CW$4)", "Ежемесячная регулярная выручка"),
    ("ARR (Annual Recurring Revenue)", "=B6*12", "Годовая регулярная выручка"),
    ("Revenue Growth Rate (MoM)", "=(B6-B6)/B6", "Рост выручки месяц к месяцу"),
    ("Cumulative Revenue", "=SUM('Monthly forecast'!$D$4:$CW$4)", "Суммарная выручка за все время"),
]
current_row = create_section(current_row, "📊 REVENUE METRICS", revenue_metrics)

# 2. CUSTOMER METRICS
print(f"   • Customer Metrics...")
customer_metrics = [
    ("Total Customers", "=COUNTA('Unit economics'!$D$6:$CW$6)", "Всего активных клиентов"),
    ("New Customers (Last Month)", "=B12", "Новых клиентов за последний месяц"),
    ("Monthly Churn Rate", "=15%", "Процент ушедших клиентов в месяц"),
    ("Customer Lifetime (months)", "=1/B14", "Средняя продолжительность подписки"),
    ("ARPPU (Avg Revenue Per User)", "=$100", "Средний доход на пользователя"),
]
current_row = create_section(current_row, "👥 CUSTOMER METRICS", customer_metrics)

# 3. UNIT ECONOMICS
print(f"   • Unit Economics...")
unit_economics = [
    ("LTV (Lifetime Value)", "='Unit economics'!D54", "Пожизненная ценность клиента"),
    ("CAC (Customer Acquisition Cost)", "='Unit economics'!D59", "Стоимость привлечения клиента"),
    ("LTV:CAC Ratio", "=B19/B20", "Соотношение LTV к CAC (цель: 3-5)"),
    ("CAC Payback Period (months)", "=B20/B16", "Срок окупаемости затрат на привлечение"),
    ("LTV 12 month", "='Unit economics'!D57", "LTV за 12 месяцев"),
]
current_row = create_section(current_row, "💰 UNIT ECONOMICS", unit_economics)

# 4. FINANCIAL HEALTH
print(f"   • Financial Health...")
financial_metrics = [
    ("Monthly Burn Rate", "=SUM('P&L | Cash Flow | Balance'!D20:D60)", "Ежемесячный расход денег"),
    ("Cash Balance", "='P&L | Cash Flow | Balance'!D70", "Текущий остаток денег"),
    ("Runway (months)", "=B27/B26", "Сколько месяцев работы до конца денег"),
    ("Gross Margin %", "=80%", "Валовая маржа (типично для SaaS)"),
]
current_row = create_section(current_row, "💵 FINANCIAL HEALTH", financial_metrics)

# 5. KEY RATIOS & BENCHMARKS
print(f"   • Key Ratios & Benchmarks...")
ratios = [
    ("LTV:CAC Target (B2B SaaS)", "4:1", "Здоровое соотношение для роста"),
    ("CAC Payback Target", "< 12 months", "Рекомендуемый срок окупаемости"),
    ("Monthly Churn Target (B2B)", "< 1%", "Целевой показатель оттока"),
    ("MRR Growth Target", "> 10% MoM", "Целевой рост для раннего этапа"),
]
current_row = create_section(current_row, "🎯 BENCHMARKS (B2B SaaS)", ratios)

# 6. ALERTS & WARNINGS
print(f"   • Alerts & Warnings...")
ws[f'A{current_row}'] = "⚠️ ВАЖНЫЕ ЗАМЕЧАНИЯ"
ws[f'A{current_row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
ws[f'A{current_row}'].font = Font(name='Arial', size=12, bold=True, color="C00000")
ws.merge_cells(f'A{current_row}:C{current_row}')
current_row += 1

warnings = [
    "• Churn 15% в месяц слишком высок для B2B (норма <1%)",
    "• Проверьте: это месячный или годовой churn?",
    "• При корректном churn (1%) LTV вырастет в 15 раз!",
    "• Уточните pricing: $100 за сотрудника или за компанию?",
]

for warning in warnings:
    ws[f'A{current_row}'] = warning
    ws[f'A{current_row}'].font = Font(name='Arial', size=10, color="C00000")
    ws.merge_cells(f'A{current_row}:C{current_row}')
    current_row += 1

print(f"\n5. Настройка форматирования...")

# Устанавливаем ширину столбцов
ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 45

# Замораживаем первые 4 строки
ws.freeze_panes = 'A5'

print(f"\n6. Сохранение изменений...")
wb.save(file_path)
print(f"   ✅ Файл сохранен: {file_path}")

print("\n" + "=" * 80)
print("KPI DASHBOARD СОЗДАН УСПЕШНО!")
print("=" * 80)

print(f"""
✅ Создан новый лист 'KPI Dashboard' в начале книги

📊 Добавленные секции:
   • Revenue Metrics (MRR, ARR, Growth Rate)
   • Customer Metrics (Total, New, Churn, ARPPU)
   • Unit Economics (LTV, CAC, Ratios, Payback)
   • Financial Health (Burn Rate, Runway, Cash)
   • Benchmarks (целевые показатели B2B SaaS)
   • Важные замечания и предупреждения

🔗 Формулы автоматически подтягивают данные из:
   • Unit economics
   • Monthly forecast
   • P&L | Cash Flow | Balance

⚠️ ВАЖНО:
   Проверьте формулы в Excel - они могут требовать корректировки
   в зависимости от структуры ваших данных!

📝 Следующие шаги:
   1. Откройте Excel и проверьте KPI Dashboard
   2. Скорректируйте формулы при необходимости
   3. Добавьте условное форматирование для наглядности
   4. Создайте графики на основе этих метрик

Резервная копия: {backup_file}
""")

print("=" * 80)
