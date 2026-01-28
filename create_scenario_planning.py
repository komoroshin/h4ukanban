from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
import shutil
from datetime import datetime

file_path = 'finmodel/Office Open XML spreadsheet.xlsx'
backup_file = f'finmodel/Office Open XML spreadsheet_before_Scenarios_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

print("=" * 80)
print("СОЗДАНИЕ СЦЕНАРНОГО ПЛАНИРОВАНИЯ")
print("=" * 80)

# Создаем резервную копию
print(f"\n1. Создание резервной копии...")
shutil.copy2(file_path, backup_file)
print(f"   ✅ Резервная копия: {backup_file}")

# Загружаем workbook
print(f"\n2. Загрузка модели...")
wb = load_workbook(file_path)

# Проверяем, существует ли уже Scenarios
if 'Scenarios' in wb.sheetnames:
    print(f"   ⚠️  Лист 'Scenarios' уже существует, будет перезаписан")
    del wb['Scenarios']

# Создаем новый лист
ws = wb.create_sheet('Scenarios', 1)  # Вторым листом после KPI Dashboard
print(f"   ✅ Создан новый лист 'Scenarios'")

print(f"\n3. Настройка стилей...")

# Стили
header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
header_font = Font(name='Arial', size=14, bold=True, color="FFFFFF")
scenario_fills = {
    'base': PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
    'best': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    'worst': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}
scenario_fonts = {
    'base': Font(name='Arial', size=11, bold=True, color="2F75B5"),
    'best': Font(name='Arial', size=11, bold=True, color="006100"),
    'worst': Font(name='Arial', size=11, bold=True, color="9C0006"),
}
metric_font = Font(name='Arial', size=10)
value_font = Font(name='Arial', size=10, bold=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

print(f"\n4. Создание структуры сценариев...")

# Заголовок
ws['A1'] = 'Healthy4U'
ws['A1'].font = Font(name='Arial', size=18, bold=True, color="2F75B5")
ws['A2'] = 'СЦЕНАРНОЕ ПЛАНИРОВАНИЕ'
ws['A2'].font = Font(name='Arial', size=16, bold=True)
ws['A3'] = 'Прогнозы развития бизнеса на 24 месяца'
ws['A3'].font = Font(name='Arial', size=10, italic=True, color="7F7F7F")

row = 5

# Описание сценариев
ws[f'A{row}'] = 'ОПИСАНИЕ СЦЕНАРИЕВ'
ws[f'A{row}'].fill = header_fill
ws[f'A{row}'].font = header_font
ws.merge_cells(f'A{row}:D{row}')
row += 1

scenarios_description = [
    ('Base Case (Базовый)', 'Наиболее вероятный сценарий развития', scenario_fills['base'], scenario_fonts['base']),
    ('Best Case (Оптимистичный)', 'Благоприятное развитие событий, быстрый рост', scenario_fills['best'], scenario_fonts['best']),
    ('Worst Case (Пессимистичный)', 'Неблагоприятные условия, медленный рост', scenario_fills['worst'], scenario_fonts['worst']),
]

for scenario_name, description, fill, font in scenarios_description:
    ws[f'A{row}'] = scenario_name
    ws[f'A{row}'].fill = fill
    ws[f'A{row}'].font = font
    ws[f'B{row}'] = description
    ws[f'B{row}'].font = metric_font
    ws.merge_cells(f'B{row}:D{row}')
    row += 1

row += 1

# Ключевые параметры сценариев
print(f"   • Параметры сценариев...")

ws[f'A{row}'] = 'КЛЮЧЕВЫЕ ПАРАМЕТРЫ СЦЕНАРИЕВ'
ws[f'A{row}'].fill = header_fill
ws[f'A{row}'].font = header_font
ws.merge_cells(f'A{row}:D{row}')
row += 1

# Заголовки таблицы
headers = ['Параметр', 'Base Case', 'Best Case', 'Worst Case']
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=row, column=col_idx)
    cell.value = header
    cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    cell.font = Font(name='Arial', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

row += 1
param_start_row = row

# Параметры
parameters = [
    ('📈 Customer Growth', 'Рост клиентской базы'),
    ('   Monthly New Customers (Month 1-6)', ['50', '80', '30']),
    ('   Monthly New Customers (Month 7-12)', ['100', '150', '60']),
    ('   Monthly New Customers (Month 13-24)', ['150', '250', '80']),
    ('', ''),
    ('💸 Churn Rate', 'Отток клиентов'),
    ('   Monthly Churn %', ['1.0%', '0.5%', '2.0%']),
    ('   Annual Churn %', ['12%', '6%', '24%']),
    ('', ''),
    ('💰 Unit Economics', 'Экономика привлечения'),
    ('   ARPPU ($/month)', ['$100', '$120', '$80']),
    ('   CAC ($)', ['$150', '$120', '$180']),
    ('   LTV:CAC Ratio', ['4.0x', '6.0x', '2.7x']),
    ('', ''),
    ('🎯 Market & Growth', 'Рынок и рост'),
    ('   Market Penetration (Year 2)', ['5%', '10%', '2%']),
    ('   Revenue Growth (MoM avg)', ['15%', '25%', '8%']),
    ('   Team Growth (by Month 24)', ['+30 people', '+50 people', '+15 people']),
    ('', ''),
    ('💵 Financial', 'Финансовые показатели'),
    ('   Gross Margin', ['80%', '85%', '75%']),
    ('   Operating Expenses Growth', ['Умеренный', 'Агрессивный', 'Консервативный']),
    ('   Break-even Month', ['Month 18', 'Month 12', 'Month 24+']),
]

for param in parameters:
    if len(param) == 2 and param[1] and not isinstance(param[1], list):
        # Категория
        ws[f'A{row}'] = param[0]
        ws[f'A{row}'].font = Font(name='Arial', size=10, bold=True, color="2F75B5")
        ws.merge_cells(f'A{row}:D{row}')
    elif len(param) == 2 and param[1] == '':
        # Пустая строка
        pass
    else:
        # Параметр со значениями
        ws[f'A{row}'] = param[0]
        ws[f'A{row}'].font = metric_font
        ws[f'A{row}'].border = border

        # Значения для каждого сценария
        values = param[1]
        for col_idx, value in enumerate(values, start=2):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = value
            cell.font = value_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

            # Цвет фона в зависимости от сценария
            if col_idx == 2:  # Base
                cell.fill = scenario_fills['base']
            elif col_idx == 3:  # Best
                cell.fill = scenario_fills['best']
            elif col_idx == 4:  # Worst
                cell.fill = scenario_fills['worst']

    row += 1

row += 1

# Прогнозные метрики
print(f"   • Прогнозные метрики...")

ws[f'A{row}'] = 'ПРОГНОЗНЫЕ МЕТРИКИ НА 24 МЕСЯЦА'
ws[f'A{row}'].fill = header_fill
ws[f'A{row}'].font = header_font
ws.merge_cells(f'A{row}:D{row}')
row += 1

# Таблица прогнозов
forecast_headers = ['Метрика', 'Base Case', 'Best Case', 'Worst Case']
for col_idx, header in enumerate(forecast_headers, start=1):
    cell = ws.cell(row=row, column=col_idx)
    cell.value = header
    cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    cell.font = Font(name='Arial', size=10, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

row += 1

# Прогнозные метрики
forecasts = [
    ('После 12 месяцев:', '', '', ''),
    ('   Total Customers', '600', '960', '360'),
    ('   MRR', '$60,000', '$115,200', '$28,800'),
    ('   ARR', '$720,000', '$1,382,400', '$345,600'),
    ('   Cumulative Revenue', '$420,000', '$720,000', '$210,000'),
    ('', '', '', ''),
    ('После 24 месяцев:', '', '', ''),
    ('   Total Customers', '2,100', '4,200', '1,200'),
    ('   MRR', '$210,000', '$504,000', '$96,000'),
    ('   ARR', '$2,520,000', '$6,048,000', '$1,152,000'),
    ('   Cumulative Revenue', '$1,890,000', '$3,780,000', '$756,000'),
    ('', '', '', ''),
    ('Ключевые вехи:', '', '', ''),
    ('   Break-even (месяц)', '18', '12', '24+'),
    ('   Profitability ARR', '$2.5M', '$1.5M', '$3.5M+'),
    ('   Funding needed', '$1.5M', '$2.5M', '$800k'),
]

for forecast in forecasts:
    if forecast[0].endswith(':'):
        # Категория
        ws[f'A{row}'] = forecast[0]
        ws[f'A{row}'].font = Font(name='Arial', size=10, bold=True, color="2F75B5")
        ws.merge_cells(f'A{row}:D{row}')
    elif forecast[1] == '':
        # Пустая строка
        pass
    else:
        # Метрика со значениями
        ws[f'A{row}'] = forecast[0]
        ws[f'A{row}'].font = metric_font
        ws[f'A{row}'].border = border

        # Значения для каждого сценария
        for col_idx, value in enumerate(forecast[1:], start=2):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = value
            cell.font = value_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

            # Цвет фона
            if col_idx == 2:  # Base
                cell.fill = scenario_fills['base']
            elif col_idx == 3:  # Best
                cell.fill = scenario_fills['best']
            elif col_idx == 4:  # Worst
                cell.fill = scenario_fills['worst']

    row += 1

row += 2

# Рекомендации
print(f"   • Рекомендации...")

ws[f'A{row}'] = '💡 РЕКОМЕНДАЦИИ ПО ИСПОЛЬЗОВАНИЮ'
ws[f'A{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True, color="C65911")
ws.merge_cells(f'A{row}:D{row}')
row += 1

recommendations = [
    "✓ Base Case - используйте для внутреннего планирования и бюджетирования",
    "✓ Best Case - покажите инвесторам потенциал при благоприятных условиях",
    "✓ Worst Case - используйте для стресс-тестирования и планирования рисков",
    "✓ Регулярно пересматривайте сценарии на основе актуальных данных",
    "✓ Фокусируйтесь на Base Case, но готовьтесь к Worst Case",
]

for rec in recommendations:
    ws[f'A{row}'] = rec
    ws[f'A{row}'].font = Font(name='Arial', size=10)
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

print(f"\n5. Настройка форматирования...")

# Устанавливаем ширину столбцов
ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 20

# Замораживаем первые 4 строки
ws.freeze_panes = 'A5'

print(f"\n6. Сохранение изменений...")
wb.save(file_path)
print(f"   ✅ Файл сохранен: {file_path}")

print("\n" + "=" * 80)
print("СЦЕНАРНОЕ ПЛАНИРОВАНИЕ СОЗДАНО УСПЕШНО!")
print("=" * 80)

print(f"""
✅ Создан новый лист 'Scenarios'

📊 Добавлено 3 сценария:
   • Base Case - наиболее вероятный сценарий
   • Best Case - оптимистичный (быстрый рост, низкий churn)
   • Worst Case - пессимистичный (медленный рост, высокий churn)

📈 Ключевые параметры:
   • Customer Growth (рост клиентов по месяцам)
   • Churn Rate (отток клиентов)
   • Unit Economics (ARPPU, CAC, LTV:CAC)
   • Market & Growth (проникновение рынка, рост выручки)
   • Financial (маржа, расходы, break-even)

🎯 Прогнозы:
   • 12 месяцев: MRR от $28k до $115k
   • 24 месяца: MRR от $96k до $504k
   • Break-even: от 12 до 24+ месяцев
   • Funding needed: от $800k до $2.5M

💡 Использование:
   • Base Case → внутреннее планирование
   • Best Case → презентации инвесторам
   • Worst Case → управление рисками

📝 Следующие шаги:
   1. Проверьте сценарии в Excel
   2. Адаптируйте параметры под ваши реалии
   3. Создайте графики для визуализации
   4. Используйте для презентаций

Резервная копия: {backup_file}
""")

print("=" * 80)
