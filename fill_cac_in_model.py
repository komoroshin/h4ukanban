from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import shutil
from datetime import datetime

file_path = 'finmodel/Office Open XML spreadsheet.xlsx'
backup_file = f'finmodel/Office Open XML spreadsheet_before_CAC_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

print("=" * 80)
print("ЗАПОЛНЕНИЕ CAC В ФИНАНСОВОЙ МОДЕЛИ")
print("=" * 80)

# Создаем резервную копию
print(f"\n1. Создание резервной копии...")
shutil.copy2(file_path, backup_file)
print(f"   ✅ Резервная копия: {backup_file}")

# Загружаем workbook
wb = load_workbook(file_path)
ws = wb['Unit economics']

print(f"\n2. Анализ структуры листа 'Unit economics'...")

# Ищем строку с CAC
cac_row = None
for row in range(1, 65):
    cell_value = ws[f'A{row}'].value
    if cell_value and 'CAC' in str(cell_value) and 'Cost of Outbound' in str(cell_value):
        cac_row = row
        print(f"   ✅ Найдена строка CAC: {row} ('{cell_value}')")
        break

if not cac_row:
    print("   ❌ Строка CAC не найдена!")
    exit(1)

print(f"\n3. Определение сценариев CAC на основе бенчмарков...")

# Из наших расчетов (LTV ~$600)
scenarios = {
    "Базовый (4:1)": 150,
    "Консервативный (3:1)": 200,
    "Агрессивный (2.5:1)": 240
}

for name, value in scenarios.items():
    print(f"   • {name}: ${value}")

print(f"\n4. Выбор сценария для применения...")
selected_scenario = "Базовый (4:1)"
selected_cac = scenarios[selected_scenario]
print(f"   ✅ Выбран: {selected_scenario} = ${selected_cac}")

print(f"\n5. Заполнение CAC в модели...")

# Определяем диапазон столбцов с данными
# Обычно данные начинаются со столбца D (4)
start_col = 4  # D
filled_count = 0

# Заполняем CAC с небольшим ростом со временем (симуляция увеличения затрат)
for col_idx in range(start_col, start_col + 100):  # До 100 месяцев
    col_letter = get_column_letter(col_idx)
    cell = ws[f'{col_letter}{cac_row}']

    # Проверяем, есть ли данные в этом столбце (смотрим на строку с датами)
    date_cell = ws[f'{col_letter}2']  # Строка 2 обычно содержит даты

    if date_cell.value:
        # Небольшое увеличение CAC со временем (инфляция, рост рынка)
        # Первые 6 месяцев - более низкий CAC (органический рост)
        # Затем постепенное увеличение
        month_offset = col_idx - start_col

        if month_offset < 6:
            # Ранние месяцы - органический рост, более низкий CAC
            cac_value = selected_cac * 0.6
        elif month_offset < 12:
            # Рост масштабирования
            cac_value = selected_cac * 0.8
        else:
            # Стабильный период
            cac_value = selected_cac

        cell.value = round(cac_value, 2)
        filled_count += 1

        if filled_count <= 12:
            print(f"      {cell.coordinate} = ${cac_value:.2f}")

print(f"\n   ✅ Заполнено ячеек: {filled_count}")

print(f"\n6. Сохранение изменений...")
wb.save(file_path)
print(f"   ✅ Файл сохранен: {file_path}")

print("\n" + "=" * 80)
print("ИТОГОВАЯ ИНФОРМАЦИЯ")
print("=" * 80)

print(f"""
✅ CAC успешно заполнен в модели!

Применённая стратегия:
• Базовый CAC: ${selected_cac} (LTV:CAC ratio = 4:1)
• Месяцы 1-6: ${selected_cac * 0.6:.2f} (органический рост)
• Месяцы 7-12: ${selected_cac * 0.8:.2f} (масштабирование)
• Месяцы 13+: ${selected_cac:.2f} (стабильный период)

Обоснование:
• Соответствует бенчмаркам B2B SaaS (LTV:CAC = 4:1)
• CAC Payback Period ≈ 1.5 месяца (при MRR $100)
• Учитывает постепенное увеличение затрат при масштабировании

⚠️ ВАЖНО:
Теперь метрики LTV/CAC будут рассчитываться автоматически!
Откройте файл в Excel и проверьте результаты.

Резервная копия: {backup_file}
""")

print("=" * 80)
