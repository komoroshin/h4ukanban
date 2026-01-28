import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import shutil
from datetime import datetime

# Путь к файлу
original_file = 'finmodel/Office Open XML spreadsheet.xlsx'
backup_file = f'finmodel/Office Open XML spreadsheet_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

print("=" * 80)
print("ИСПРАВЛЕНИЕ ОШИБОК В UNIT ECONOMICS")
print("=" * 80)

# Создаем резервную копию
print(f"\n1. Создание резервной копии...")
shutil.copy2(original_file, backup_file)
print(f"   ✅ Резервная копия создана: {backup_file}")

# Загружаем workbook
print(f"\n2. Загрузка файла...")
wb = load_workbook(original_file)
ws = wb['Unit economics']
print(f"   ✅ Лист 'Unit economics' загружен")

# Находим проблемные строки
# Строка 60 (Excel) = индекс 60 (openpyxl использует 1-based индексацию)
problem_rows = {
    60: "LTV/CAC",
    61: "LTV 3 month/CAC",
    62: "LTV 6 month/CAC",
    63: "LTV 12 month/CAC"
}

print(f"\n3. Исправление формул с делением на ноль...")
fixed_count = 0

for row_num, row_name in problem_rows.items():
    print(f"\n   Обработка строки {row_num}: {row_name}")

    # Проходим по всем столбцам (начиная с D = 4)
    for col_num in range(4, ws.max_column + 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f'{col_letter}{row_num}']

        # Проверяем, есть ли формула
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            old_formula = cell.value

            # Заменяем формулу на защищенную версию
            # Например: =D54/D$59 -> =IFERROR(D54/D$59, "-")
            # Или используем IF: =IF(D$59=0, "-", D54/D$59)

            # Вариант с IFERROR (более чистый)
            new_formula = f'=IFERROR({old_formula[1:]}, "-")'

            cell.value = new_formula
            fixed_count += 1

            if col_num <= 6:  # Показываем первые несколько для примера
                print(f"      {cell.coordinate}: {old_formula} -> {new_formula}")

print(f"\n   ✅ Исправлено формул: {fixed_count}")

# Сохраняем изменения
print(f"\n4. Сохранение изменений...")
wb.save(original_file)
print(f"   ✅ Файл сохранен: {original_file}")

print("\n" + "=" * 80)
print("ИСПРАВЛЕНИЯ ПРИМЕНЕНЫ УСПЕШНО")
print("=" * 80)
print(f"\nРезервная копия: {backup_file}")
print(f"Обновленный файл: {original_file}")
print(f"\nИсправлено ячеек: {fixed_count}")
print("\nТеперь вместо ошибок #DIV/0! будет отображаться '-'")
print("Когда вы заполните данные CAC, метрики будут рассчитываться автоматически.")
print("=" * 80)
