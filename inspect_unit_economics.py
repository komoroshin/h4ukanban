import pandas as pd
from openpyxl import load_workbook

file_path = 'finmodel/Office Open XML spreadsheet.xlsx'

# Читаем лист Unit economics
df = pd.read_excel(file_path, sheet_name='Unit economics', header=None)

print("=" * 80)
print("ДЕТАЛЬНЫЙ АНАЛИЗ ЛИСТА 'UNIT ECONOMICS'")
print("=" * 80)

# Показываем структуру
print(f"\nРазмер: {df.shape[0]} строк x {df.shape[1]} столбцов")

# Показываем первые строки для понимания заголовков
print("\nПервые 10 строк (первые 10 столбцов):")
print(df.iloc[:10, :10].to_string())

# Показываем строки 55-65 где ошибки (первые 10 столбцов)
print("\n" + "=" * 80)
print("ПРОБЛЕМНАЯ ОБЛАСТЬ - строки 55-65:")
print("=" * 80)
print(df.iloc[55:65, :10].to_string())

# Загружаем workbook для проверки формул
wb = load_workbook(file_path)
ws = wb['Unit economics']

print("\n" + "=" * 80)
print("ФОРМУЛЫ В ПРОБЛЕМНЫХ ЯЧЕЙКАХ (строка 60):")
print("=" * 80)

# Проверяем формулы в строке 60 (индекс 61 в Excel, 60 в Python 0-based)
for col_idx, col_letter in enumerate(['D', 'E', 'F', 'G', 'H'], start=4):
    cell = ws[f'{col_letter}60']
    print(f"\nЯчейка {cell.coordinate}:")
    print(f"  Значение: {cell.value}")
    if hasattr(cell, 'formula') or cell.data_type == 'f':
        # Пытаемся получить формулу
        print(f"  Тип: {cell.data_type}")

# Показываем метки строк (столбец A или B)
print("\n" + "=" * 80)
print("МЕТКИ СТРОК (55-65):")
print("=" * 80)
for idx in range(55, 65):
    label_a = df.iloc[idx, 0] if idx < len(df) else None
    label_b = df.iloc[idx, 1] if idx < len(df) and len(df.columns) > 1 else None
    print(f"Строка {idx+1}: A='{label_a}' | B='{label_b}'")

print("\n" + "=" * 80)
