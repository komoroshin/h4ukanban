#!/usr/bin/env python3
"""
Healthy4U ROI Calculator
Generate ROI analysis for AI implementation in medical clinics

Usage:
    python roi_calculator.py --physicians 3 --hourly-rate 50 --hours-saved 4

Output:
    - Excel file with detailed ROI calculations
    - PDF summary report (optional)
"""

import argparse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import sys


class ROICalculator:
    """Calculate ROI for Healthy4U platform implementation"""

    # Pricing constants (from PRICING_STRATEGY_UPDATED.md)
    PRICING = {
        'standard': 249,  # $/month per physician
        'founders_circle': 174,  # 30% discount
        'early_adopter': 199,  # 20% discount
        'launch_partner': 224,  # 10% discount
    }

    # Cost assumptions
    IMPLEMENTATION_COSTS = {
        'tier1_diy': 0,  # Free with subscription
        'tier2_guided': 5000,  # Guided implementation
        'tier3_full': 15000,  # Full-service (midpoint of $10K-20K)
    }

    def __init__(self, num_physicians, hourly_rate, hours_saved_per_week,
                 pricing_tier='founders_circle', implementation_tier='tier1_diy',
                 current_costs=0):
        """
        Initialize ROI calculator

        Args:
            num_physicians: Number of physicians in practice
            hourly_rate: Physician hourly rate ($/hour)
            hours_saved_per_week: Hours saved per physician per week
            pricing_tier: Healthy4U pricing tier (founders_circle, early_adopter, etc.)
            implementation_tier: Implementation tier (tier1_diy, tier2_guided, tier3_full)
            current_costs: Current monthly costs for existing solutions ($/month)
        """
        self.num_physicians = num_physicians
        self.hourly_rate = hourly_rate
        self.hours_saved_per_week = hours_saved_per_week
        self.pricing_tier = pricing_tier
        self.implementation_tier = implementation_tier
        self.current_costs = current_costs

    def calculate_monthly_savings(self):
        """Calculate monthly time savings in dollars"""
        # Assume 4 weeks per month
        monthly_hours_saved = self.hours_saved_per_week * self.num_physicians * 4
        monthly_savings = monthly_hours_saved * self.hourly_rate
        return monthly_savings

    def calculate_healthy4u_costs(self):
        """Calculate Healthy4U monthly and implementation costs"""
        monthly_cost = self.PRICING[self.pricing_tier] * self.num_physicians
        implementation_cost = self.IMPLEMENTATION_COSTS[self.implementation_tier]
        return monthly_cost, implementation_cost

    def calculate_net_savings(self):
        """Calculate net monthly savings (savings - new costs + old costs eliminated)"""
        monthly_savings = self.calculate_monthly_savings()
        monthly_cost, _ = self.calculate_healthy4u_costs()
        net_monthly_savings = monthly_savings - monthly_cost + self.current_costs
        return net_monthly_savings

    def calculate_payback_period(self):
        """Calculate payback period in months"""
        net_monthly_savings = self.calculate_net_savings()
        _, implementation_cost = self.calculate_healthy4u_costs()

        if net_monthly_savings <= 0:
            return float('inf')  # Never pays back

        payback_months = implementation_cost / net_monthly_savings
        return payback_months

    def calculate_3year_roi(self):
        """Calculate 3-year total ROI"""
        net_monthly_savings = self.calculate_net_savings()
        _, implementation_cost = self.calculate_healthy4u_costs()

        total_3year_savings = net_monthly_savings * 36
        total_investment = implementation_cost

        roi_dollars = total_3year_savings - total_investment
        roi_percent = (roi_dollars / total_investment * 100) if total_investment > 0 else float('inf')

        return roi_dollars, roi_percent

    def generate_excel_report(self, filename='roi_analysis.xlsx'):
        """Generate detailed Excel report"""
        wb = openpyxl.Workbook()

        # Create sheets
        ws_summary = wb.active
        ws_summary.title = "ROI Summary"
        ws_details = wb.create_sheet("Detailed Analysis")
        ws_comparison = wb.create_sheet("Competitor Comparison")

        # Generate each sheet
        self._create_summary_sheet(ws_summary)
        self._create_details_sheet(ws_details)
        self._create_comparison_sheet(ws_comparison)

        # Save workbook
        wb.save(filename)
        print(f"✅ ROI report generated: {filename}")
        return filename

    def _create_summary_sheet(self, ws):
        """Create executive summary sheet"""
        # Header
        ws['A1'] = "Healthy4U ROI Analysis"
        ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.merge_cells('A1:D1')

        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A2'].font = Font(size=10, italic=True)
        ws.merge_cells('A2:D2')

        # Input assumptions
        row = 4
        ws[f'A{row}'] = "INPUT ASSUMPTIONS"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 1

        inputs = [
            ("Number of Physicians", self.num_physicians, "physicians"),
            ("Physician Hourly Rate", f"${self.hourly_rate}", "/hour"),
            ("Hours Saved per Week (per physician)", self.hours_saved_per_week, "hours"),
            ("Healthy4U Pricing Tier", self.pricing_tier.replace('_', ' ').title(), ""),
            ("Implementation Tier", self.implementation_tier.replace('_', ' ').title(), ""),
            ("Current Monthly Costs", f"${self.current_costs:,.0f}", "(to be eliminated)"),
        ]

        for label, value, unit in inputs:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'C{row}'] = value
            ws[f'D{row}'] = unit
            row += 1

        # Key Results
        row += 2
        ws[f'A{row}'] = "KEY RESULTS"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        # Calculate results
        monthly_savings = self.calculate_monthly_savings()
        monthly_cost, implementation_cost = self.calculate_healthy4u_costs()
        net_monthly_savings = self.calculate_net_savings()
        payback_months = self.calculate_payback_period()
        roi_3yr_dollars, roi_3yr_percent = self.calculate_3year_roi()

        results = [
            ("💰 Monthly Time Savings", f"${monthly_savings:,.0f}", "/month", "green"),
            ("💳 Healthy4U Monthly Cost", f"${monthly_cost:,.0f}", "/month", "red"),
            ("✅ Net Monthly Savings", f"${net_monthly_savings:,.0f}", "/month", "green" if net_monthly_savings > 0 else "red"),
            ("", "", "", None),
            ("📊 Implementation Cost (One-Time)", f"${implementation_cost:,.0f}", "", "orange"),
            ("⏱️ Payback Period", f"{payback_months:.1f}" if payback_months != float('inf') else "Never", "months", "orange"),
            ("", "", "", None),
            ("🎯 Year 1 Net Savings", f"${net_monthly_savings * 12 - implementation_cost:,.0f}", "", "green" if (net_monthly_savings * 12 - implementation_cost) > 0 else "red"),
            ("🚀 3-Year ROI", f"${roi_3yr_dollars:,.0f}", f"({roi_3yr_percent:.0f}%)", "green" if roi_3yr_dollars > 0 else "red"),
        ]

        for label, value, unit, color in results:
            if label == "":
                row += 1
                continue

            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'C{row}'] = value
            ws[f'C{row}'].font = Font(size=12, bold=True)

            if color == "green":
                ws[f'C{row}'].font = Font(size=12, bold=True, color="006100")
                ws[f'C{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif color == "red":
                ws[f'C{row}'].font = Font(size=12, bold=True, color="9C0006")
                ws[f'C{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif color == "orange":
                ws[f'C{row}'].font = Font(size=12, bold=True, color="9C5700")
                ws[f'C{row}'].fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

            ws[f'D{row}'] = unit
            row += 1

        # Recommendation
        row += 2
        ws[f'A{row}'] = "💡 RECOMMENDATION"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        row += 1

        if payback_months <= 6:
            recommendation = "✅ STRONG BUSINESS CASE: Payback in under 6 months. Highly recommended."
            color = "C6EFCE"
        elif payback_months <= 12:
            recommendation = "✅ GOOD BUSINESS CASE: Payback within first year. Recommended."
            color = "FFEB9C"
        elif payback_months <= 24:
            recommendation = "⚠️ MODERATE BUSINESS CASE: Payback in 1-2 years. Consider if strategic priority."
            color = "FFEB9C"
        else:
            recommendation = "❌ WEAK BUSINESS CASE: Payback >2 years. Reconsider assumptions or wait."
            color = "FFC7CE"

        ws[f'A{row}'] = recommendation
        ws[f'A{row}'].font = Font(size=11, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.merge_cells(f'A{row}:D{row}')

        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15

    def _create_details_sheet(self, ws):
        """Create detailed monthly breakdown sheet"""
        # Header
        ws['A1'] = "36-Month Detailed Projection"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')

        # Column headers
        headers = ["Month", "Cumulative Savings", "Cumulative Cost", "Net Benefit", "Payback Status"]
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Calculate month-by-month
        net_monthly_savings = self.calculate_net_savings()
        _, implementation_cost = self.calculate_healthy4u_costs()

        payback_reached = False

        for month in range(1, 37):
            row = month + 3

            cumulative_savings = net_monthly_savings * month
            cumulative_cost = implementation_cost  # One-time cost
            net_benefit = cumulative_savings - cumulative_cost

            if net_benefit >= 0 and not payback_reached:
                payback_status = "✅ PAYBACK REACHED"
                payback_reached = True
            elif payback_reached:
                payback_status = "✅ Positive ROI"
            else:
                payback_status = "⏳ Not yet"

            ws[f'A{row}'] = month
            ws[f'B{row}'] = f"${cumulative_savings:,.0f}"
            ws[f'C{row}'] = f"${cumulative_cost:,.0f}"
            ws[f'D{row}'] = f"${net_benefit:,.0f}"
            ws[f'E{row}'] = payback_status

            # Color coding
            if net_benefit >= 0:
                ws[f'D{row}'].font = Font(color="006100")
                ws[f'D{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                ws[f'D{row}'].font = Font(color="9C0006")

        # Adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 20

    def _create_comparison_sheet(self, ws):
        """Create competitor comparison sheet"""
        # Header
        ws['A1'] = "Healthy4U vs Competitors"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:G1')

        ws['A2'] = "Annual Cost Comparison"
        ws['A2'].font = Font(size=12, italic=True)
        ws.merge_cells('A2:G2')

        # Competitor data (from PRICING_MARKET_ANALYSIS_2026.md)
        competitors = [
            {
                'name': 'Healthy4U',
                'monthly_per_physician': self.PRICING[self.pricing_tier],
                'features': 'Scribe + AI Consultant + Patient App',
                'patient_side': 'Yes',
                'integration': 'Medium',
                'implementation': self.IMPLEMENTATION_COSTS[self.implementation_tier]
            },
            {
                'name': 'Freed.ai',
                'monthly_per_physician': 99,
                'features': 'Scribe only',
                'patient_side': 'No',
                'integration': 'Easy',
                'implementation': 0
            },
            {
                'name': 'Suki Assistant',
                'monthly_per_physician': 399,
                'features': 'Scribe + Voice',
                'patient_side': 'No',
                'integration': 'Medium',
                'implementation': 0
            },
            {
                'name': 'DeepScribe',
                'monthly_per_physician': 750,
                'features': 'Premium Scribe + Coding',
                'patient_side': 'No',
                'integration': 'Complex',
                'implementation': 5000
            },
            {
                'name': 'Custom Build',
                'monthly_per_physician': 0,  # No recurring
                'features': 'Fully Customizable',
                'patient_side': 'Maybe',
                'integration': 'Perfect',
                'implementation': 75000  # $50K-100K average
            },
        ]

        # Table headers
        headers = ["Solution", "Monthly/Physician", "Annual Cost", "Features", "Patient Side", "Integration", "Total Year 1"]
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Populate data
        for idx, comp in enumerate(competitors, start=5):
            annual_cost = comp['monthly_per_physician'] * self.num_physicians * 12
            total_year1 = annual_cost + comp['implementation']

            ws[f'A{idx}'] = comp['name']
            ws[f'A{idx}'].font = Font(bold=True) if comp['name'] == 'Healthy4U' else Font()

            ws[f'B{idx}'] = f"${comp['monthly_per_physician']}"
            ws[f'C{idx}'] = f"${annual_cost:,.0f}"
            ws[f'D{idx}'] = comp['features']
            ws[f'E{idx}'] = comp['patient_side']
            ws[f'F{idx}'] = comp['integration']
            ws[f'G{idx}'] = f"${total_year1:,.0f}"

            # Highlight Healthy4U row
            if comp['name'] == 'Healthy4U':
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                    ws[f'{col}{idx}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 15


def main():
    parser = argparse.ArgumentParser(description='Healthy4U ROI Calculator')
    parser.add_argument('--physicians', type=int, required=True, help='Number of physicians')
    parser.add_argument('--hourly-rate', type=float, required=True, help='Physician hourly rate ($/hour)')
    parser.add_argument('--hours-saved', type=float, required=True, help='Hours saved per physician per week')
    parser.add_argument('--pricing-tier', default='founders_circle',
                        choices=['standard', 'founders_circle', 'early_adopter', 'launch_partner'],
                        help='Healthy4U pricing tier')
    parser.add_argument('--implementation', default='tier1_diy',
                        choices=['tier1_diy', 'tier2_guided', 'tier3_full'],
                        help='Implementation tier')
    parser.add_argument('--current-costs', type=float, default=0,
                        help='Current monthly costs to be eliminated ($/month)')
    parser.add_argument('--output', default='roi_analysis.xlsx', help='Output filename')

    args = parser.parse_args()

    # Create calculator
    calc = ROICalculator(
        num_physicians=args.physicians,
        hourly_rate=args.hourly_rate,
        hours_saved_per_week=args.hours_saved,
        pricing_tier=args.pricing_tier,
        implementation_tier=args.implementation,
        current_costs=args.current_costs
    )

    # Generate report
    print(f"\n🔢 Calculating ROI for {args.physicians} physicians...")
    calc.generate_excel_report(args.output)

    # Print summary to console
    print("\n" + "="*60)
    print("ROI SUMMARY")
    print("="*60)

    monthly_savings = calc.calculate_monthly_savings()
    monthly_cost, implementation_cost = calc.calculate_healthy4u_costs()
    net_monthly_savings = calc.calculate_net_savings()
    payback_months = calc.calculate_payback_period()
    roi_3yr_dollars, roi_3yr_percent = calc.calculate_3year_roi()

    print(f"Monthly Time Savings: ${monthly_savings:,.0f}/month")
    print(f"Healthy4U Monthly Cost: ${monthly_cost:,.0f}/month")
    print(f"Net Monthly Savings: ${net_monthly_savings:,.0f}/month")
    print(f"Implementation Cost: ${implementation_cost:,.0f} (one-time)")
    print(f"Payback Period: {payback_months:.1f} months" if payback_months != float('inf') else "Payback Period: Never")
    print(f"3-Year ROI: ${roi_3yr_dollars:,.0f} ({roi_3yr_percent:.0f}%)")
    print("="*60 + "\n")


if __name__ == '__main__':
    main()
